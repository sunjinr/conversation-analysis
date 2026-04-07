#!/usr/bin/env python3
"""
skisight_analysis 数据引擎

只提供三个核心能力：
1. load_sessions() — 加载内嵌的 JSON 会话数据
2. keyword_filter() — 关键词粗筛 + 日期提取
3. generate_report() — 接收 LLM 分类结果，生成三 Sheet Excel

分类由 Qoder 大模型完成，本脚本不参与。
"""

import json
import re
import os
import time
import requests
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell


# ─── LLM API 配置 ───────────────────────────────────────────────

LLM_API_KEY = "sk-sp-fd7cd74769e04ff78641600b5339886e"
LLM_BASE_URL = "https://coding.dashscope.aliyuncs.com/v1"
LLM_MODEL = "qwen3-coder-plus"  # 可用的编程模型
LLM_TIMEOUT = 120
LLM_MAX_RETRIES = 3
LLM_RATE_LIMIT_DELAY = 2  # 批量调用间延迟（秒）


# ─── 默认分析维度（含优化建议）───────────────────────────────────

DEFAULT_DIMENSIONS = {
    "协同效率低下": {
        "definition": "涉及多个部门或商家时，响应慢，互相推诿",
        "suggestion": "建立跨部门快速响应机制，设定SLA时限；客服需主动跟进并同步进度给用户，避免用户被动等待",
    },
    "交互设计反人类": {
        "definition": "表单、按钮设计不合理，让用户操作很费劲",
        "suggestion": "简化用户操作路径，减少表单填写和材料上传步骤；优化移动端按钮布局和操作流程",
    },
    "话术模板化与无效回复": {
        "definition": "回复全是标准套话，没有针对性，或者全是安抚情绪但没有实际解决方案",
        "suggestion": "优化机器人话术库，增加场景化回复；引入动态回复机制，根据用户问题上下文生成针对性回答",
    },
    "解释不清与信息冗余": {
        "definition": "解释得云里雾里用户听不懂，或者废话太多重点不突出",
        "suggestion": "推行'结论先行'回复规范，先给答案再解释原因；限制单次回复长度，使用结构化格式",
    },
    "前后矛盾与口径不一": {
        "definition": "不同的客服或机器人与人给的答案不一样，导致用户不知道信谁的",
        "suggestion": "建立统一知识库和标准答案库；机器人和人工使用同一套话术和规则，定期校准",
    },
    "意图识别失败（含多意图/小语种）": {
        "definition": "未能识别出用户的真实需求，包括多意图只识别一个，或小语种方言导致误判",
        "suggestion": "加强多语言NLP能力；支持多意图识别和拆分处理；增加意图确认环节",
    },
    "复杂/模糊意图无法处理": {
        "definition": "遇到稍微复杂的问题，系统无法进行多线程判断，强行简化为通用答案",
        "suggestion": "引入追问机制，当用户意图不清晰时主动提问；复杂场景自动升级至人工",
    },
    "情绪与风险识别缺失": {
        "definition": "无法识别用户是否愤怒、焦虑，或是否涉及自杀、曝光等高风险事件",
        "suggestion": "部署情绪识别模型，当检测到愤怒/焦虑时自动调整回复策略；高风险事件立即触发人工介入",
    },
    "能力局限（只能查不能办）": {
        "definition": "机器人只能做简单的查询和引导，一旦涉及修改、操作、复杂判断就必须转人工",
        "suggestion": "扩展机器人操作能力，支持常见操作类需求（修改地址、取消订单等）；减少转人工场景",
    },
    "承诺未兑现与反复无进展": {
        "definition": "承诺了能解决结果没解决，或者一直说在处理但永远没结果",
        "suggestion": "建立工单追踪系统，承诺事项自动生成跟进任务；超时未解决自动提醒和升级",
    },
    "缺乏补救与补偿": {
        "definition": "因平台或服务方过错导致用户受损，却没有提供任何补偿或补救措施",
        "suggestion": "制定标准化补偿方案，在确认平台责任后自动触发补偿流程；客服拥有补偿决策权限",
    },
    "转人工障碍（入口隐藏/不肯转）": {
        "definition": "用户明确想转人工但系统不给入口，或入口藏得很深",
        "suggestion": "在机器人对话界面固定位置放置转人工入口；用户连续2次表达转人工意向时自动转接",
    },
    "转接后的信息丢失": {
        "definition": "从机器人转到人工后，人工客服不知道前面发生了什么，导致用户需要重复描述问题",
        "suggestion": "建立机器人与人工之间的会话上下文自动传递机制；人工接起时自动展示历史对话摘要和已收集信息",
    },
}


# ─── 数据加载 ───────────────────────────────────────────────────

def load_sessions():
    """加载会话数据。优先从环境变量指定的路径加载，fallback 到内嵌的 chat_data.json"""
    # 优先使用环境变量指定的数据文件路径
    data_path = os.environ.get('SESSIONS_DATA_PATH')
    if data_path and os.path.exists(data_path):
        with open(data_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    # fallback 到脚本目录下的 chat_data.json
    data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chat_data.json')
    if os.path.exists(data_path):
        with open(data_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    # 最后 fallback：返回空列表
    return []


# 关键词扩展映射，补充 LLM 可能遗漏的相关词
KEYWORD_EXPANSIONS = {
    '支付': ['支付', '付款', 'pay', 'payment', '扣款', '充值', '转账', '支付宝', 'wechat pay'],
    '退款': ['退款', 'refund', '退钱', '退回', '返款', '退货退款'],
    '物流': ['物流', '快递', '配送', '运输', 'shipping', 'delivery', '包裹', '签收'],
    '发票': ['发票', 'invoice', '开票', '报销', '税务'],
    '供应商': ['供应商', '商家', 'seller', 'vendor', 'supplier', '纠纷', '争议'],
    '邮箱': ['邮箱', 'email', '邮件', 'mailbox', '修改邮箱', '换邮箱'],
    '账户': ['账户', '账号', 'account', '登录', '密码', '注册', '注销'],
    '优惠券': ['优惠券', 'coupon', '折扣', 'promo', '代金券', '红包'],
}


def expand_keywords(keywords):
    """扩展关键词"""
    expanded = set(keywords)
    for kw in keywords:
        if kw in KEYWORD_EXPANSIONS:
            expanded.update(KEYWORD_EXPANSIONS[kw])
        # 也做反向匹配
        for parent, children in KEYWORD_EXPANSIONS.items():
            if kw.lower() in [c.lower() for c in children]:
                expanded.add(parent)
                expanded.update(children)
    return list(expanded)


def _extract_date(messages):
    """从消息列表中提取日期"""
    for msg in messages:
        m = re.search(r'\[(\d{4}-\d{2}-\d{2})', msg)
        if m:
            return m.group(1)
    return None


def _check_human_transfer(human_messages):
    """检查是否转人工：human_messages 中是否有客服发言"""
    for msg in human_messages:
        if '客服:' in msg and '机器人:' not in msg and '命令消息' not in msg:
            return True
    return False


def keyword_filter(sessions, keywords):
    """
    关键词粗筛：过滤出包含关键词的不满意会话。
    返回 (filtered_sessions, session_map)
    session_map: {session_id: session_raw_data} 用于后续获取原始内容
    """
    if not keywords:
        # 无关键词则返回所有不满意会话
        filtered = [s for s in sessions if s.get('unsatisfied')]
    else:
        expanded = expand_keywords(keywords)
        filtered = []
        for s in sessions:
            if not s.get('unsatisfied'):
                continue
            all_text = ' '.join(s.get('bot_messages', []) + s.get('human_messages', [])).lower()
            if any(kw.lower() in all_text for kw in expanded):
                filtered.append(s)

    # 构建 session_map
    session_map = {s.get('session_id', ''): s for s in sessions if s.get('session_id')}

    return filtered, session_map


# ─── LLM API 调用 ───────────────────────────────────────────────

def call_llm(prompt, system_prompt=None, max_retries=None):
    """调用 LLM API，返回响应文本"""
    if max_retries is None:
        max_retries = LLM_MAX_RETRIES

    headers = {
        'Authorization': f'Bearer {LLM_API_KEY}',
        'Content-Type': 'application/json',
    }

    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    payload = {
        "model": LLM_MODEL,
        "messages": messages,
        "temperature": 0.3,
        "max_tokens": 4096,
    }

    for attempt in range(max_retries):
        try:
            resp = requests.post(
                f"{LLM_BASE_URL}/chat/completions",
                headers=headers,
                json=payload,
                timeout=LLM_TIMEOUT,
            )
            resp.raise_for_status()
            data = resp.json()
            return data['choices'][0]['message']['content'].strip()
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(LLM_RATE_LIMIT_DELAY * (attempt + 1))
            else:
                raise RuntimeError(f"LLM API 调用失败 ({max_retries} 次): {e}")


def llm_refine_topic(sessions, topic_label):
    """
    LLM 语义精筛：逐条判断会话是否与主题相关。
    返回 [{'session_id': xxx, 'relevant': True/False, 'reason': '...'}]
    """
    if not sessions or not topic_label:
        return sessions

    judgments = []
    batch_size = 10

    for i in range(0, len(sessions), batch_size):
        batch = sessions[i:i + batch_size]
        items_text = ""
        for idx, s in enumerate(batch):
            bot_chat = '\n'.join(s.get('bot_messages', []))[:2000]
            human_chat = '\n'.join(s.get('human_messages', []))[:2000]
            items_text += f"\n--- 会话 {idx+1} (session_id: {s.get('session_id', '')}) ---\n"
            items_text += f"机器人对话:\n{bot_chat}\n"
            items_text += f"人工对话:\n{human_chat}\n"

        system_prompt = (
            "你是一个客服对话分析专家。请判断以下每条会话是否与指定主题相关。\n"
            "判断标准：只要会话中涉及或提及了该主题相关的内容，即视为相关，不要求该主题必须是会话的核心问题。\n"
            "例如：主题是'支付失败'，如果会话中提到了支付问题、扣款异常、付款不成功等内容，即使主要在讨论退款，也应判断为相关。\n"
            "请严格按照 JSON 格式输出，不要输出其他内容。\n"
            "输出格式: [{\"index\": 1, \"relevant\": true/false, \"reason\": \"简要原因\"}, ...]"
        )

        prompt = (
            f"主题: {topic_label}\n"
            f"请判断以下 {len(batch)} 条会话是否涉及该主题（宽松匹配，只要涉及即算相关）：\n"
            f"{items_text}\n"
            f"请输出 JSON 数组，每个元素包含 index（从1开始）、relevant（布尔值）、reason（简要原因）。"
        )

        try:
            result = call_llm(prompt, system_prompt)
            # 解析 JSON
            result = re.sub(r'^```(?:json)?\s*', '', result)
            result = re.sub(r'\s*```$', '', result)
            judgments_list = json.loads(result)

            for j in judgments_list:
                idx = j.get('index', 0) - 1
                if 0 <= idx < len(batch):
                    judgments.append({
                        'session_id': batch[idx].get('session_id', ''),
                        'relevant': j.get('relevant', False),
                        'reason': j.get('reason', ''),
                    })
        except Exception as e:
            # 出错时保守处理：全部保留
            for s in batch:
                judgments.append({
                    'session_id': s.get('session_id', ''),
                    'relevant': True,
                    'reason': f'LLM 判断出错: {e}',
                })

        if i + batch_size < len(sessions):
            time.sleep(LLM_RATE_LIMIT_DELAY)

    return judgments


def llm_classify_sessions(sessions, dimensions=None):
    """
    LLM 逐条分类：对每条会话判断不满原因属于哪些维度。
    返回 [{'session_id': xxx, 'primary_dimension': xxx, 'secondary_dimensions': [...], 'detail': '...', 'new_dimension': None}]
    """
    if dimensions is None:
        dimensions = DEFAULT_DIMENSIONS

    results = []
    batch_size = 5

    dim_list = list(dimensions.keys())
    dim_defs = "\n".join([f"- {d}: {dimensions[d]['definition']}" for d in dim_list])

    for i in range(0, len(sessions), batch_size):
        batch = sessions[i:i + batch_size]
        items_text = ""
        for idx, s in enumerate(batch):
            bot_chat = '\n'.join(s.get('bot_messages', []))[:2000]
            human_chat = '\n'.join(s.get('human_messages', []))[:2000]
            raw_reason = s.get('raw_reason', s.get('raw_unsatisfied_reason', ''))
            items_text += f"\n--- 会话 {idx+1} (session_id: {s.get('session_id', '')}) ---\n"
            items_text += f"原始不满意原因: {raw_reason}\n"
            items_text += f"机器人对话:\n{bot_chat}\n"
            items_text += f"人工对话:\n{human_chat}\n"

        system_prompt = (
            "你是一个客服对话分析专家。请分析以下每条会话中用户不满的原因，并归类到指定维度。\n"
            "可选维度：\n" + dim_defs + "\n\n"
            "如果现有维度都不匹配，可以创建新维度（设置 new_dimension 字段）。\n"
            "请严格按照 JSON 格式输出，不要输出其他内容。\n"
            '输出格式: [{"index": 1, "primary_dimension": "维度名", "secondary_dimensions": ["次要维度1"], '
            '"detail": "具体原因说明", "new_dimension": null, "new_dimension_definition": ""}, ...]'
        )

        prompt = (
            f"请分析以下 {len(batch)} 条会话的不满原因：\n"
            f"{items_text}\n"
            f"请输出 JSON 数组，每个元素包含 index（从1开始）、primary_dimension（主维度）、"
            f"secondary_dimensions（次维度数组，可为空）、detail（具体原因说明）、"
            f"new_dimension（新维度名，如无则为null）、new_dimension_definition（新维度定义）。"
        )

        try:
            result = call_llm(prompt, system_prompt)
            result = re.sub(r'^```(?:json)?\s*', '', result)
            result = re.sub(r'\s*```$', '', result)
            results_list = json.loads(result)

            for r in results_list:
                idx = r.get('index', 0) - 1
                if 0 <= idx < len(batch):
                    results.append({
                        'session_id': batch[idx].get('session_id', ''),
                        'primary_dimension': r.get('primary_dimension', '其他'),
                        'secondary_dimensions': r.get('secondary_dimensions', []),
                        'detail': r.get('detail', ''),
                        'new_dimension': r.get('new_dimension'),
                        'new_dimension_definition': r.get('new_dimension_definition', ''),
                    })
        except Exception as e:
            for s in batch:
                results.append({
                    'session_id': s.get('session_id', ''),
                    'primary_dimension': '其他',
                    'secondary_dimensions': [],
                    'detail': f'LLM 分类出错: {e}',
                    'new_dimension': None,
                    'new_dimension_definition': '',
                })

        if i + batch_size < len(sessions):
            time.sleep(LLM_RATE_LIMIT_DELAY)

    return results


# ─── 辅助函数 ────────────────────────────────────────────────────

def refine_topic_with_llm(unsat_sessions, topic_label, session_map=None):
    """
    LLM 精筛辅助：包装 llm_refine_topic，返回精筛结果。
    unsat_sessions 可以是原始 session 列表或 session_id 列表。
    """
    if not topic_label or not unsat_sessions:
        return unsat_sessions, []

    # 如果传入的是 session_id 列表，转换为完整 session 数据
    if isinstance(unsat_sessions[0], str) and session_map:
        sessions_to_refine = [session_map[sid] for sid in unsat_sessions if sid in session_map]
    else:
        sessions_to_refine = unsat_sessions

    judgments = llm_refine_topic(sessions_to_refine, topic_label)
    relevant_ids = {j['session_id'] for j in judgments if j.get('relevant', False)}

    # 返回精筛后的 session_id 列表和 judgments
    filtered = [s['session_id'] if isinstance(s, dict) else s for s in unsat_sessions
                if (s['session_id'] if isinstance(s, dict) else s) in relevant_ids]

    return filtered, judgments


def apply_llm_refinement(unsat_sessions, llm_judgments, session_map=None):
    """
    应用 LLM 精筛结果，过滤出相关会话。
    llm_judgments: [{'session_id': xxx, 'relevant': True/False, 'reason': '...'}]
    返回过滤后的会话列表。
    """
    if not llm_judgments:
        return unsat_sessions

    relevant_ids = {j['session_id'] for j in llm_judgments if j.get('relevant', False)}

    result = []
    for s in unsat_sessions:
        sid = s['session_id'] if isinstance(s, dict) else s
        if sid in relevant_ids:
            if isinstance(s, dict):
                result.append(s)
            elif session_map and sid in session_map:
                result.append(session_map[sid])

    return result


def extract_keywords_from_question(question):
    """
    从用户问题中提取关键词（简易版本，实际由 LLM 意图解析完成）。
    """
    # 简单关键词提取：匹配已知关键词
    all_keywords = set()
    for parent, children in KEYWORD_EXPANSIONS.items():
        for kw in children:
            if kw.lower() in question.lower():
                all_keywords.add(kw)
                all_keywords.add(parent)
    return list(all_keywords)


def extract_topic_label(keywords):
    """根据关键词推断主题标签"""
    for kw in keywords:
        if kw in KEYWORD_EXPANSIONS:
            return kw
    return keywords[0] if keywords else '全部'


def parse_user_intent(question, dimensions=None):
    """
    LLM 意图解析：从用户问题中提取结构化参数。
    返回 {'topic_label': str, 'keywords': [...], 'time_range': str, 'custom_dimensions': dict|None, 'user_questions': [...], 'analysis_type': str}
    """
    system_prompt = (
        "你是一个客服对话分析系统的意图解析引擎。请从用户的问题中提取结构化参数。\n"
        "输出严格的 JSON 格式，包含以下字段：\n"
        "- topic_label: 主题标签（如'支付失败'、'退款问题'）\n"
        "- keywords: 相关关键词列表\n"
        "- time_range: 时间范围（如'202603'、'2026-03-15'）\n"
        "- custom_dimensions: 自定义维度定义（如果用户指定了，否则为null）\n"
        "- user_questions: 用户的具体问题列表\n"
        "- analysis_type: 'topic_analysis'（主题分析）或 'trend_analysis'（趋势分析）"
    )

    prompt = f"请解析以下用户问题的意图：\n{question}"

    try:
        result = call_llm(prompt, system_prompt)
        result = re.sub(r'^```(?:json)?\s*', '', result)
        result = re.sub(r'\s*```$', '', result)
        return json.loads(result)
    except Exception as e:
        # 回退到简易解析
        keywords = extract_keywords_from_question(question)
        return {
            'topic_label': extract_topic_label(keywords),
            'keywords': keywords,
            'time_range': '',
            'custom_dimensions': None,
            'user_questions': [question],
            'analysis_type': 'topic_analysis',
        }


# ─── 完整分析流程 ────────────────────────────────────────────────

def run_full_analysis(user_question, topic_keywords=None, topic_label=None,
                      custom_dimensions=None, output_path=None,
                      date_from=None, date_to=None):
    """
    完整分析流程：
    1. LLM 意图解析（如果未提供参数）
    2. 加载数据（可按日期过滤）
    3. 关键词粗筛
    4. LLM 语义精筛
    5. LLM 逐条分类
    6. LLM 回答用户问题
    7. 生成 Excel 报告
    """
    # Step 0: 意图解析
    intent = None
    if not topic_keywords or not topic_label:
        intent = parse_user_intent(user_question, custom_dimensions)
        if not topic_keywords:
            topic_keywords = intent.get('keywords', [])
        if not topic_label:
            topic_label = intent.get('topic_label', '全部')
        user_questions = intent.get('user_questions', [user_question])
    else:
        user_questions = [user_question]

    # 打印主题信息供前端解析
    print(f"  主题: {topic_label}")
    print(f"  关键词: {topic_keywords}")

    # 使用自定义维度或默认维度
    dimensions = custom_dimensions if custom_dimensions else DEFAULT_DIMENSIONS

    # Step 1: 加载数据
    sessions = load_sessions()
    print(f"  ✓ 加载 {len(sessions)} 条会话")

    # 按日期范围过滤
    if date_from or date_to:
        before_count = len(sessions)
        filtered = []
        for s in sessions:
            all_msgs = s.get('bot_messages', []) + s.get('human_messages', [])
            session_date = _extract_date(all_msgs)
            if not session_date:
                continue
            if date_from and session_date < date_from:
                continue
            if date_to and session_date > date_to:
                continue
            filtered.append(s)
        sessions = filtered
        print(f"  ✓ 日期过滤 ({date_from or '...'} ~ {date_to or '...'}): {before_count} → {len(sessions)} 条")

    # Step 2a: 关键词粗筛
    filtered_sessions, session_map = keyword_filter(sessions, topic_keywords)
    print(f"  ✓ 关键词匹配: {len(filtered_sessions)} 条会话")

    # Step 2b: LLM 语义精筛
    relevant_sessions, llm_judgments = refine_topic_with_llm(
        filtered_sessions, topic_label, session_map
    )
    print(f"  ✓ 精筛完成: {len(relevant_sessions)} 条相关会话")

    # 获取精筛后的完整 session 数据
    if isinstance(relevant_sessions[0], str) if relevant_sessions else False:
        refined_sessions = [session_map[sid] for sid in relevant_sessions if sid in session_map]
    else:
        refined_sessions = relevant_sessions

    # Step 3: LLM 逐条分类
    llm_results = llm_classify_sessions(refined_sessions, dimensions)
    print(f"  ✓ 分类完成: {len(llm_results)} 条会话")

    # 注册新维度
    for r in llm_results:
        new_dim = r.get('new_dimension')
        if new_dim and new_dim not in dimensions:
            dimensions[new_dim] = {
                'definition': r.get('new_dimension_definition', ''),
                'suggestion': f"针对「{new_dim}」建议进一步分析并制定优化方案",
            }

    # Step 4 & 5: 生成报告
    report_path = generate_report(
        sessions=sessions,
        session_map=session_map,
        llm_results=llm_results,
        llm_judgments=llm_judgments,
        topic_label=topic_label,
        topic_keywords=topic_keywords,
        user_questions=user_questions,
        dimensions=dimensions,
        output_path=output_path,
    )

    return report_path


# ─── 用户问题回答（动态数据处理） ──────────────────────────────────


def _filter_records_by_date_range(records, start, end):
    """按日期范围过滤记录，start/end 为 YYYY-MM-DD 或 YYYY-MM 前缀"""
    return [r for r in records
            if r.get('date', '') and r['date'] >= start and (r['date'] <= end or r['date'].startswith(end))]


def _group_records_by_period(records, group_by='month'):
    """按天/月分桶，返回 OrderedDict"""
    from collections import OrderedDict
    groups = {}
    for r in records:
        d = r.get('date', '')
        if not d:
            continue
        key = d[:7] if group_by == 'month' else d
        groups.setdefault(key, []).append(r)
    return OrderedDict(sorted(groups.items()))


def _parse_question_intent(question, available_dates, available_dimensions):
    """LLM 解析用户问题意图 → 结构化 JSON；失败时正则 fallback"""
    date_range_str = f"{min(available_dates)} ~ {max(available_dates)}" if available_dates else "无"
    dims_str = '、'.join(available_dimensions[:10]) if available_dimensions else "无"

    system_prompt = (
        "你是一个数据分析意图解析器。分析用户问题，返回JSON。\n"
        "analysis_type 可选值：comparison / trend / root_cause / distribution / general\n"
        f"数据日期范围：{date_range_str}\n"
        f"数据维度：{dims_str}\n\n"
        "严格按JSON格式返回，不要其他内容：\n"
        '{"analysis_type":"comparison","date_ranges":[{"label":"3月","prefix":"2026-03"},{"label":"4月","prefix":"2026-04"}],'
        '"dimensions_focus":null,"group_by":"month"}'
    )

    try:
        raw = call_llm(question, system_prompt, max_retries=1)
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            intent = json.loads(json_match.group())
            if 'analysis_type' in intent:
                return intent
    except Exception:
        pass
    return _parse_intent_fallback(question, available_dates)


def _parse_intent_fallback(question, available_dates):
    """正则 fallback 解析意图"""
    q = question.lower()
    intent = {"analysis_type": "general", "date_ranges": [], "dimensions_focus": None, "group_by": "month"}

    if any(kw in q for kw in ['对比', '比较', 'vs', '相比', '变多', '变少']):
        intent['analysis_type'] = 'comparison'
    elif any(kw in q for kw in ['趋势', '走势']):
        intent['analysis_type'] = 'trend'
    elif any(kw in q for kw in ['原因', '为什么', '根因', '主要']):
        intent['analysis_type'] = 'root_cause'
    elif any(kw in q for kw in ['分布', '占比', '比例', '构成']):
        intent['analysis_type'] = 'distribution'

    # 提取日期
    date_matches = re.findall(r'(\d{4})[\-年]?(\d{1,2})月?', question)
    if not date_matches:
        month_matches = re.findall(r'(\d{1,2})月', question)
        if month_matches and available_dates:
            year = available_dates[0][:4]
            date_matches = [(year, m) for m in month_matches]
    for m in date_matches:
        if len(m) == 2:
            prefix = f"{m[0]}-{int(m[1]):02d}"
            intent['date_ranges'].append({"label": f"{int(m[1])}月", "prefix": prefix})
    return intent


def _dispatch_analysis(intent, records):
    """根据意图路由到对应计算函数"""
    t = intent.get('analysis_type', 'general')
    dispatch = {
        'comparison': _compute_comparison,
        'trend': _compute_trend,
        'root_cause': _compute_root_cause,
        'distribution': _compute_distribution,
    }
    fn = dispatch.get(t, _compute_general)
    return fn(records, intent)


def _compute_comparison(records, intent):
    """按时段对比：各时段维度分布 + 差值"""
    date_ranges = intent.get('date_ranges', [])
    if len(date_ranges) < 2:
        return _compute_general(records, intent)

    periods = []
    for dr in date_ranges:
        prefix = dr.get('prefix', '')
        label = dr.get('label', prefix)
        recs = [r for r in records if r.get('date', '').startswith(prefix)]
        dim_counts = Counter(r.get('primary_dimension', '其他') for r in recs)
        human = sum(1 for r in recs if r.get('has_human'))
        details = []
        for r in recs[:3]:
            d = r.get('detail', '') or r.get('raw_reason', '')
            if d:
                details.append(d[:150])
        periods.append({
            "label": label, "prefix": prefix, "total": len(recs),
            "dim_counts": dict(dim_counts.most_common()),
            "human_count": human,
            "human_rate": f"{human/len(recs)*100:.1f}%" if recs else "0%",
            "sample_details": details,
        })

    deltas = []
    for i in range(len(periods) - 1):
        p1, p2 = periods[i], periods[i + 1]
        count_change = p2['total'] - p1['total']
        pct = (count_change / p1['total'] * 100) if p1['total'] else 0
        all_dims = set(list(p1['dim_counts'].keys()) + list(p2['dim_counts'].keys()))
        dim_changes = []
        for dim in sorted(all_dims):
            c1, c2 = p1['dim_counts'].get(dim, 0), p2['dim_counts'].get(dim, 0)
            if c1 != c2:
                dim_changes.append({"dim": dim, "from": c1, "to": c2, "change": c2 - c1})
        dim_changes.sort(key=lambda x: abs(x['change']), reverse=True)
        deltas.append({
            "from_label": p1['label'], "to_label": p2['label'],
            "count_change": count_change, "count_change_pct": f"{pct:+.1f}%",
            "dim_changes": dim_changes,
        })
    return {"analysis_type": "comparison", "periods": periods, "deltas": deltas}


def _compute_trend(records, intent):
    """时间序列趋势"""
    group_by = intent.get('group_by', 'month')
    grouped = _group_records_by_period(records, group_by)
    time_series = []
    for period, recs in grouped.items():
        dim_counts = Counter(r.get('primary_dimension', '其他') for r in recs)
        time_series.append({"period": period, "total": len(recs), "dimensions": dict(dim_counts.most_common(5))})

    trend = "数据点不足"
    if len(time_series) >= 2:
        first, last = time_series[0]['total'], time_series[-1]['total']
        trend = "上升" if last > first * 1.1 else ("下降" if last < first * 0.9 else "基本持平")

    peak = max(time_series, key=lambda x: x['total']) if time_series else None
    trough = min(time_series, key=lambda x: x['total']) if time_series else None
    return {
        "analysis_type": "trend", "time_series": time_series, "overall_trend": trend,
        "peak": peak['period'] if peak else None, "trough": trough['period'] if trough else None,
    }


def _compute_root_cause(records, intent):
    """原因深度分析"""
    dim_counts = Counter(r.get('primary_dimension', '其他') for r in records)
    total = len(records)
    ranking = []
    for dim, cnt in dim_counts.most_common():
        dim_recs = [r for r in records if r.get('primary_dimension') == dim]
        human = sum(1 for r in dim_recs if r.get('has_human'))
        seen, cases = set(), []
        for r in dim_recs:
            d = r.get('detail', '') or r.get('raw_reason', '')
            if d and d not in seen:
                seen.add(d)
                cases.append(d[:200])
                if len(cases) >= 3:
                    break
        sec_counts = Counter()
        for r in dim_recs:
            secs = r.get('secondary_dimensions', '')
            if secs and secs != '-':
                for s in re.split(r'[、,，]', secs):
                    s = s.strip()
                    if s:
                        sec_counts[s] += 1
        ranking.append({
            "dimension": dim, "count": cnt, "percentage": f"{cnt/total*100:.1f}%",
            "human_rate": f"{human/len(dim_recs)*100:.1f}%" if dim_recs else "0%",
            "cases": cases, "secondary_co_occurrence": dict(sec_counts.most_common(3)),
        })
    return {"analysis_type": "root_cause", "total": total, "ranking": ranking}


def _compute_distribution(records, intent):
    """分布统计 + 交叉表"""
    total = len(records)
    dim_counts = Counter(r.get('primary_dimension', '其他') for r in records)
    date_counts = Counter(r.get('date', '') for r in records if r.get('date'))
    human_total = sum(1 for r in records if r.get('has_human'))
    grouped = _group_records_by_period(records, 'month')
    cross_tab = {}
    for period, recs in grouped.items():
        cross_tab[period] = dict(Counter(r.get('primary_dimension', '其他') for r in recs).most_common())
    return {
        "analysis_type": "distribution", "total": total,
        "dim_distribution": [{"dim": d, "count": c, "pct": f"{c/total*100:.1f}%"} for d, c in dim_counts.most_common()],
        "date_distribution": [{"date": d, "count": c} for d, c in sorted(date_counts.items())],
        "human_transfer": {"total": human_total, "rate": f"{human_total/total*100:.1f}%" if total else "0%"},
        "cross_tab": cross_tab,
    }


def _compute_general(records, intent):
    """通用综合摘要"""
    total = len(records)
    dim_counts = Counter(r.get('primary_dimension', '其他') for r in records)
    date_counts = Counter(r.get('date', '') for r in records if r.get('date'))
    human_total = sum(1 for r in records if r.get('has_human'))
    dim_details = {}
    for dim, _ in dim_counts.most_common():
        dim_recs = [r for r in records if r.get('primary_dimension') == dim]
        details = []
        for r in dim_recs[:3]:
            d = r.get('detail', '') or r.get('raw_reason', '')
            if d:
                details.append(d[:150])
        dim_details[dim] = details
    return {
        "analysis_type": "general", "total": total,
        "dim_counts": dict(dim_counts.most_common()),
        "date_counts": dict(sorted(date_counts.items())),
        "human_rate": f"{human_total/total*100:.1f}%" if total else "0%",
        "dim_details": dim_details,
    }


def _format_computed_results(computed_data, topic_label=""):
    """将计算结果格式化为 LLM 易读的结构化文本"""
    t = computed_data.get('analysis_type', 'general')
    lines = []
    if topic_label:
        lines.append(f"分析主题: {topic_label}")

    if t == 'comparison':
        lines.append("\n【各时段数据】")
        for p in computed_data.get('periods', []):
            lines.append(f"\n{p['label']}（共{p['total']}条，转人工{p['human_rate']}）:")
            for dim, cnt in p['dim_counts'].items():
                pct = cnt / p['total'] * 100 if p['total'] else 0
                lines.append(f"  {dim}: {cnt}次({pct:.1f}%)")
            if p.get('sample_details'):
                lines.append("  典型案例:")
                for d in p['sample_details']:
                    lines.append(f"    - {d}")
        lines.append("\n【对比变化】")
        for delta in computed_data.get('deltas', []):
            lines.append(f"{delta['from_label']} → {delta['to_label']}: 总量变化 {delta['count_change']:+d} ({delta['count_change_pct']})")
            for dc in delta['dim_changes'][:5]:
                lines.append(f"  {dc['dim']}: {dc['from']} → {dc['to']} ({dc['change']:+d})")

    elif t == 'trend':
        lines.append(f"\n【整体趋势】: {computed_data.get('overall_trend', '未知')}")
        if computed_data.get('peak'):
            lines.append(f"峰值: {computed_data['peak']}")
        if computed_data.get('trough'):
            lines.append(f"谷值: {computed_data['trough']}")
        lines.append("\n【时间序列】")
        for ts in computed_data.get('time_series', []):
            top_dims = ', '.join(f"{d}({c})" for d, c in list(ts['dimensions'].items())[:3])
            lines.append(f"  {ts['period']}: {ts['total']}条 | {top_dims}")

    elif t == 'root_cause':
        lines.append(f"\n总记录: {computed_data.get('total', 0)}条")
        lines.append("\n【原因排名】")
        for r in computed_data.get('ranking', []):
            lines.append(f"\n{r['dimension']}（{r['count']}次，{r['percentage']}，转人工{r['human_rate']}）:")
            for c in r.get('cases', []):
                lines.append(f"  案例: {c}")
            if r.get('secondary_co_occurrence'):
                co = ', '.join(f"{k}({v})" for k, v in r['secondary_co_occurrence'].items())
                lines.append(f"  关联维度: {co}")

    elif t == 'distribution':
        lines.append(f"\n总记录: {computed_data.get('total', 0)}条")
        lines.append(f"转人工率: {computed_data.get('human_transfer', {}).get('rate', '0%')}")
        lines.append("\n【维度分布】")
        for item in computed_data.get('dim_distribution', []):
            lines.append(f"  {item['dim']}: {item['count']}次({item['pct']})")
        cross = computed_data.get('cross_tab', {})
        if len(cross) > 1:
            lines.append("\n【月份×维度交叉表】")
            for period, dims in cross.items():
                top = ', '.join(f"{d}({c})" for d, c in list(dims.items())[:5])
                lines.append(f"  {period}: {top}")

    else:  # general
        lines.append(f"\n总记录: {computed_data.get('total', 0)}条，转人工率: {computed_data.get('human_rate', '0%')}")
        lines.append("\n【维度分布】")
        for dim, cnt in computed_data.get('dim_counts', {}).items():
            total = computed_data.get('total', 1)
            lines.append(f"  {dim}: {cnt}次({cnt/total*100:.1f}%)")
        lines.append("\n【典型案例】")
        for dim, details in computed_data.get('dim_details', {}).items():
            if details:
                lines.append(f"  [{dim}]")
                for d in details:
                    lines.append(f"    - {d}")

    result = '\n'.join(lines)
    if len(result) > 2500:
        result = result[:2500] + "\n...（数据已截断）"
    return result


def _generate_answer_with_llm(question, analysis_type, formatted_results, topic_label=""):
    """基于计算结果让 LLM 生成自然语言答案"""
    topic_desc = f"「{topic_label}」" if topic_label else "相关场景"
    type_prompts = {
        "comparison": f"你是数据分析专家。以下是{topic_desc}不同时期的对比计算结果，请基于精确数据回答，重点说明变化方向、幅度和驱动因素。",
        "trend": f"你是数据分析专家。以下是{topic_desc}的趋势数据，请描述趋势方向、拐点和主要变化维度。",
        "root_cause": f"你是数据分析专家。以下是{topic_desc}的不满原因深度分析，请分析主要原因并引用具体案例。",
        "distribution": f"你是数据分析专家。以下是{topic_desc}的统计分布数据，请描述分布特征和重点。",
        "general": f"你是数据分析专家。以下是{topic_desc}的综合分析数据，请直接回答问题。",
    }
    system_prompt = type_prompts.get(analysis_type, type_prompts["general"])
    system_prompt += "回答要简洁有条理，使用具体数字，不要说无法找到数据。"

    prompt = f"以下是基于实际数据计算的分析结果：\n{formatted_results}\n\n用户问题：{question}\n\n请直接回答。"
    return call_llm(prompt, system_prompt, max_retries=2)


def _answer_user_question_with_llm(question, analysis_records, sessions_data,
                                    topic_label="", topic_keywords=None):
    """动态数据处理问答：解析意图 → Python计算 → LLM生成答案"""
    if not analysis_records:
        return "数据量不足，无法回答。"
    try:
        available_dates = sorted(set(r.get('date', '') for r in analysis_records if r.get('date')))
        available_dimensions = [d for d, _ in Counter(
            r.get('primary_dimension', '其他') for r in analysis_records).most_common()]

        intent = _parse_question_intent(question, available_dates, available_dimensions)
        computed = _dispatch_analysis(intent, analysis_records)
        formatted = _format_computed_results(computed, topic_label)
        answer = _generate_answer_with_llm(question, intent.get('analysis_type', 'general'),
                                            formatted, topic_label)
        return answer
    except Exception:
        return _answer_user_question_fallback(question, analysis_records)


def _answer_user_question_fallback(question, analysis_records):
    """增强 fallback：正则意图 + compute 计算 + 直接格式化"""
    if not analysis_records:
        return "数据量不足，无法回答。"
    available_dates = sorted(set(r.get('date', '') for r in analysis_records if r.get('date')))
    try:
        intent = _parse_intent_fallback(question, available_dates)
        computed = _dispatch_analysis(intent, analysis_records)
        return _format_computed_results(computed)
    except Exception:
        dim_counts = Counter(r.get('primary_dimension', '其他') for r in analysis_records)
        top = [f"{d}（{c}次，{c/len(analysis_records)*100:.1f}%）" for d, c in dim_counts.most_common(3)]
        return f"不满意主要原因：{'；'.join(top)}"


# ─── 报告生成 ────────────────────────────────────────────────────

def generate_report(sessions, session_map, llm_results, llm_judgments=None,
                    topic_label="", topic_keywords=None, user_questions=None,
                    dimensions=None, output_path=None):
    """
    生成三 Sheet Excel 报告。

    sessions: 原始会话列表
    session_map: {session_id: session_data}
    llm_results: LLM 分类结果列表
    llm_judgments: LLM 精筛结果（可选）
    """
    if dimensions is None:
        dimensions = DEFAULT_DIMENSIONS

    if not output_path:
        tag = topic_keywords[0] if topic_keywords else "全量"
        output_path = f"{tag}_不满原因分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 构建分析记录
    records = []
    seen = set()

    for item in llm_results:
        sid = item['session_id']
        if sid in seen:
            continue
        seen.add(sid)

        # 从 session_map 获取原始数据
        raw_session = session_map.get(sid, {})
        bot_messages = raw_session.get('bot_messages', [])
        human_messages = raw_session.get('human_messages', [])
        all_msgs = bot_messages + human_messages
        date_val = _extract_date(all_msgs)
        has_human_val = _check_human_transfer(human_messages)
        raw_reason = raw_session.get('raw_reason', raw_session.get('raw_unsatisfied_reason', '(未提供)'))

        primary = item.get('primary_dimension', '其他')
        secondary = item.get('secondary_dimensions', [])
        if isinstance(secondary, str):
            secondary = [s.strip() for s in re.split(r'[,，、]', secondary) if s.strip()]

        detail = item.get('detail', '') or raw_reason or '待补充'

        # 拼接客户对话内容
        bot_chat_text = '\n'.join(bot_messages) if bot_messages else '(无)'
        human_chat_text = '\n'.join(human_messages) if human_messages else '(无)'

        records.append({
            'session_id': sid,
            'user_id': raw_session.get('user_id', ''),
            'ocs_session_id': raw_session.get('ocs_session_id', ''),
            'date': date_val,
            'has_human': has_human_val,
            'raw_reason': raw_reason or '(未提供具体原因)',
            'bot_chat': bot_chat_text,
            'human_chat': human_chat_text,
            'dissatisfaction_info': raw_session.get('dissatisfaction_info', ''),
            'primary_dimension': primary,
            'secondary_dimensions': '、'.join(secondary) if secondary else '-',
            'detail': detail,
        })

    analyzed = len(records)

    # 计算统计数据
    if topic_keywords:
        expanded = expand_keywords(topic_keywords)
        mask = []
        for s in sessions:
            all_text = ' '.join(s.get('bot_messages', []) + s.get('human_messages', [])).lower()
            mask.append(any(kw.lower() in all_text for kw in expanded))

        topic_sessions = [s for i, s in enumerate(sessions) if mask[i]]
        total = len(topic_sessions)
        unsat_total = len([s for s in topic_sessions if s.get('unsatisfied')])
        human_total = len([s for s in topic_sessions if _check_human_transfer(s.get('human_messages', []))])

        if llm_judgments:
            relevant_ids = {j['session_id'] for j in llm_judgments if j.get('relevant', False)}
            total = len(relevant_ids)
            unsat_total = len([j for j in llm_judgments if j.get('relevant', False)])
            relevant_sessions_in_map = [session_map[sid] for sid in relevant_ids if sid in session_map]
            human_total = len([s for s in relevant_sessions_in_map if _check_human_transfer(s.get('human_messages', []))])
    else:
        total = len(sessions)
        unsat_total = len([s for s in sessions if s.get('unsatisfied')])
        human_total = len([s for s in sessions if _check_human_transfer(s.get('human_messages', []))])

    # 日期范围
    all_dates = [r['date'] for r in records if r.get('date')]
    date_range = f"{min(all_dates)} 至 {max(all_dates)}" if all_dates else "N/A"

    # 回答用户问题
    answered_questions = []
    if user_questions and analyzed > 0:
        for q in user_questions:
            q_text = q.get('question', str(q)) if isinstance(q, dict) else str(q)
            answer = _answer_user_question_with_llm(q_text, records, sessions,
                                                     topic_label=topic_label,
                                                     topic_keywords=topic_keywords)
            answered_questions.append({
                'question': q_text,
                'answer': answer,
            })

    # 生成 Excel
    wb = _build_excel_wb(records, total, unsat_total, human_total, date_range,
                         topic_label, answered_questions, session_map)
    wb.save(output_path)
    return output_path


def _get_dim_info(dim, dimensions=None):
    """获取维度的 definition 和 suggestion"""
    if dimensions is None:
        dimensions = DEFAULT_DIMENSIONS
    if dim in dimensions:
        info = dimensions[dim]
        return info.get('definition', ''), info.get('suggestion', '')
    return '', "建议进一步分析具体原因并制定针对性优化方案"


def _gen_findings(records):
    """生成关键发现"""
    findings = []
    n = len(records)
    if n == 0:
        return ["数据量不足，无法生成发现"]

    dc = Counter(r['primary_dimension'] for r in records)
    findings.append(f"最主要不满原因为「{dc.most_common(1)[0][0]}」，占 {dc.most_common(1)[0][1]/n*100:.1f}%（{dc.most_common(1)[0][1]}/{n}）")

    hu = [r for r in records if r.get('has_human')]
    if len(hu) > 0:
        findings.append(f"不满意会话中 {len(hu)}/{n} ({len(hu)/n*100:.1f}%) 已转人工，智能客服未能解决问题")

    dates = [r.get('date') for r in records if r.get('date')]
    if dates:
        date_counts = Counter(dates)
        sorted_dates = sorted(date_counts.items())
        if len(sorted_dates) >= 2:
            first_date, first_cnt = sorted_dates[0]
            last_date, last_cnt = sorted_dates[-1]
            if last_cnt > first_cnt:
                findings.append(f"不满意量上升：{first_date} 的 {first_cnt} -> {last_date} 的 {last_cnt}")
            elif last_cnt < first_cnt:
                findings.append(f"不满意量下降：{first_date} 的 {first_cnt} -> {last_date} 的 {last_cnt}")
            else:
                findings.append(f"不满意量持平：{first_date} 和 {last_date} 均为 {first_cnt}")

    if len(dc) >= 3:
        top3_sum = sum(cnt for _, cnt in dc.most_common(3))
        findings.append(f"前3大不满原因合计占比 {top3_sum/n*100:.1f}%，建议优先解决")

    return findings


def _build_excel_wb(records, total, unsat_total, human_total, date_range,
                    topic_label, answered_questions, session_map=None):
    """
    构建并返回 Workbook（三 Sheet）。
    session_map 用于在分析明细中获取原始会话内容。
    """
    wb = Workbook()
    analyzed = len(records)

    tf = Font(name='Microsoft YaHei', bold=True, size=16, color='1F4E79')
    stf = Font(name='Microsoft YaHei', bold=True, size=12, color='1F4E79')
    nf = Font(name='Microsoft YaHei', size=10)
    hf = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1F4E79')
    afill = PatternFill('solid', fgColor='F2F2F2')
    hlfill = PatternFill('solid', fgColor='FFF2CC')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    def set_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = tb

    def set_data(ws, row, max_col, alt=False):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = nf
            if c == 1:
                cell.font = Font(name='Microsoft YaHei', bold=True, size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = tb
            if alt:
                cell.fill = afill

    def auto_w(ws):
        for col_cells in ws.columns:
            ml = 0
            cl = None
            for cell in col_cells:
                if cl is None:
                    try:
                        cl = cell.column_letter
                    except AttributeError:
                        continue
                if cell.value and not isinstance(cell, MergedCell):
                    w = sum(2 if '\u4e00' <= ch <= '\u9fff' else 1 for ch in str(cell.value))
                    ml = max(ml, w)
            if cl and ml > 0:
                ws.column_dimensions[cl].width = min(max(ml + 3, 12), 50)

    # ═══ Sheet 1: 分析概述 ═══
    ws1 = wb.active
    ws1.title = "分析概述"
    topic_title = f"「{topic_label}」" if topic_label else ""
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f'{topic_title}客服对话不满原因分析报告'
    ws1['A1'].font = tf

    ws1['A3'] = '整体数据概览'
    ws1['A3'].font = stf

    overview = [
        ['指标', '数值'],
        ['相关会话总数', total],
        ['不满意会话数', unsat_total],
        ['不满意率', f'{unsat_total/total*100:.1f}%' if total > 0 else 'N/A'],
        ['转人工会话数', human_total],
        ['转人工率', f'{human_total/total*100:.1f}%' if total > 0 else 'N/A'],
        ['已分析不满意会话数', analyzed],
        ['分析时间段', date_range],
    ]

    for r, rd in enumerate(overview, 4):
        for c, v in enumerate(rd, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = tb
            if r == 4:
                set_header(ws1, r, 2)
            else:
                set_data(ws1, r, 2, alt=(r % 2 == 0))

    # 不满原因 TOP
    row = 4 + len(overview) + 1
    ws1.cell(row=row, column=1, value='不满原因分布（按维度）').font = stf

    if analyzed > 0:
        dim_counts = Counter(r['primary_dimension'] for r in records)
        hdrs = ['排名', '不满原因维度', '会话数', '占比']
        r = row + 1
        for c, h in enumerate(hdrs, 1):
            ws1.cell(row=r, column=c, value=h)
        set_header(ws1, r, 4)

        for rank, (dim, cnt) in enumerate(dim_counts.most_common(), 1):
            r += 1
            ws1.cell(row=r, column=1, value=rank)
            ws1.cell(row=r, column=2, value=dim)
            ws1.cell(row=r, column=3, value=cnt)
            ws1.cell(row=r, column=4, value=f'{cnt/analyzed*100:.1f}%')
            set_data(ws1, r, 4, alt=(rank % 2 == 0))

        # 用户问题回答
        if answered_questions:
            r += 2
            ws1.cell(row=r, column=1, value='用户问题回答').font = stf
            for aq in answered_questions:
                r += 1
                ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws1.cell(row=r, column=1, value=f"Q: {aq['question']}")
                ws1.cell(row=r, column=1).font = Font(name='Microsoft YaHei', bold=True, size=10)
                r += 1
                ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws1.cell(row=r, column=1, value=f"A: {aq['answer']}")
                ws1.cell(row=r, column=1).font = nf
                ws1.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

        # 关键发现
        r += 2
        ws1.cell(row=r, column=1, value='关键发现').font = stf
        findings = _gen_findings(records)
        for i, f in enumerate(findings, 1):
            r += 1
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws1.cell(row=r, column=1, value=f"{i}. {f}")
            ws1.cell(row=r, column=1).font = nf
            ws1.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
            ws1.cell(row=r, column=1).fill = hlfill

    auto_w(ws1)

    # ═══ Sheet 2: 分析详情 ═══
    ws2 = wb.create_sheet("分析详情")
    ws2.merge_cells('A1:G1')
    ws2['A1'] = '各维度详细分析与优化建议'
    ws2['A1'].font = tf

    if analyzed > 0:
        dim_counts = Counter(r['primary_dimension'] for r in records)
        hdrs = ['维度名称', '不满意会话数', '占比', '趋势', '典型案例', '优化建议', '关键特征']
        r = 2
        for c, h in enumerate(hdrs, 1):
            ws2.cell(row=r, column=c, value=h)
        set_header(ws2, r, 7)

        for idx, (dim, cnt) in enumerate(dim_counts.most_common(), 1):
            r += 1
            pct = f'{cnt/analyzed*100:.1f}%'

            trend = '-'
            dim_records = [rec for rec in records if rec['primary_dimension'] == dim]
            dim_dates = [rec['date'] for rec in dim_records if rec.get('date')]
            if dim_dates:
                date_counts = Counter(dim_dates)
                sorted_dates = sorted(date_counts.items())
                if len(sorted_dates) >= 2:
                    trend = f"{sorted_dates[-2][0]}: {sorted_dates[-2][1]} -> {sorted_dates[-1][0]}: {sorted_dates[-1][1]}"

            sample = dim_records[0]
            sid_p = sample['session_id'][:16] if sample['session_id'] else 'N/A'
            detail_p = sample['detail'][:60]
            case = f"Session: {sid_p}... | {detail_p}"

            _, suggestion = _get_dim_info(dim)
            definition, _ = _get_dim_info(dim)

            vals = [dim, cnt, pct, trend, case, suggestion, definition]
            for c, v in enumerate(vals, 1):
                ws2.cell(row=r, column=c, value=v)
            set_data(ws2, r, 7, alt=(idx % 2 == 0))
            for cc in [5, 6, 4]:
                ws2.cell(row=r, column=cc).alignment = Alignment(wrap_text=True, vertical='center')

        auto_w(ws2)

    # ═══ Sheet 3: 分析明细 ═══
    ws3 = wb.create_sheet("分析明细")
    ws3.merge_cells('A1:L1')
    ws3['A1'] = '会话级分析明细'
    ws3['A1'].font = tf

    if analyzed > 0:
        hdrs = ['序号', 'session_id', 'user_id', 'ocs_session_id',
                '客户与智能客服对话', '客户与人工客服对话',
                '客户对对话服务点击不满意', '日期',
                '主维度', '次维度', '具体原因说明']
        r = 2
        for c, h in enumerate(hdrs, 1):
            ws3.cell(row=r, column=c, value=h)
        set_header(ws3, r, 11)

        for idx, rd in enumerate(records, 1):
            r += 1

            # 从 session_map 获取完整原始会话内容
            sid = rd['session_id']
            raw_session = session_map.get(sid, {}) if session_map else {}
            bot_chat = rd.get('bot_chat', '\n'.join(raw_session.get('bot_messages', [])))
            human_chat = rd.get('human_chat', '\n'.join(raw_session.get('human_messages', [])))
            dissat_info = rd.get('dissatisfaction_info', raw_session.get('dissatisfaction_info', rd['raw_reason']))

            vals = [
                idx,
                sid,
                rd.get('user_id', raw_session.get('user_id', '')),
                rd.get('ocs_session_id', raw_session.get('ocs_session_id', '')),
                bot_chat,
                human_chat,
                dissat_info,
                rd['date'],
                rd['primary_dimension'],
                rd['secondary_dimensions'],
                rd['detail'],
            ]
            for c, v in enumerate(vals, 1):
                ws3.cell(row=r, column=c, value=v)
            set_data(ws3, r, 11, alt=(idx % 2 == 0))
            # 长文本列设置自动换行
            for cc in [5, 6, 7, 11]:
                ws3.cell(row=r, column=cc).alignment = Alignment(wrap_text=True, vertical='top')

        auto_w(ws3)

    return wb


# ─── 主入口 ─────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        question = ' '.join(sys.argv[1:])
    else:
        question = "分析所有不满意会话的原因"

    print(f"开始分析: {question}")
    print("Step 0: LLM 意图解析...")
    intent = parse_user_intent(question)
    print(f"  主题: {intent.get('topic_label')}")
    print(f"  关键词: {intent.get('keywords')}")
    print(f"  问题: {intent.get('user_questions')}")

    print("\nStep 1-5: 执行完整分析流程...")
    report_path = run_full_analysis(question)
    print(f"\n分析报告已生成: {report_path}")
